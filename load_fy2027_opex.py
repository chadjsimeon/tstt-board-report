"""
Run this script once with TSTT_Board_Data.xlsx closed in Excel.
It appends 84 FY2027 OPEX plan rows (Apr-26 to Mar-27) derived from
OPEX_AOP_FY2027.xlsx into the OPEX sheet, leaving existing rows intact.
"""

import pandas as pd
from openpyxl import load_workbook

BOARD_DATA = "TSTT_Board_Data.xlsx"
AOP_FILE   = "OPEX_AOP_FY2027.xlsx"
MONTH_ORDER = [
    "Apr-26","May-26","Jun-26","Jul-26","Aug-26","Sep-26",
    "Oct-26","Nov-26","Dec-26","Jan-27","Feb-27","Mar-27",
]

# ── Build FY2027 plan rows ─────────────────────────────────────────────────
aop = pd.read_excel(AOP_FILE, sheet_name="OPEX AOP")

plan = (
    aop.groupby(["Group", "Period"], sort=False)["Amount"]
    .sum()
    .reset_index()
    .rename(columns={"Group": "Category", "Period": "Month", "Amount": "Plan"})
)
plan["Actual"]       = 0.0
plan["Variance"]     = plan["Actual"] - plan["Plan"]
plan["Variance_Pct"] = 0.0

# Sort by the correct FY month order
plan["Month"] = pd.Categorical(plan["Month"], categories=MONTH_ORDER, ordered=True)
plan = plan.sort_values(["Month", "Category"]).reset_index(drop=True)

# ── Read existing OPEX sheet to check for duplicates and get column order ──
existing = pd.read_excel(BOARD_DATA, sheet_name="OPEX")
print(f"Existing OPEX rows: {len(existing)}")

fy27_months = set(MONTH_ORDER)
already_loaded = existing[existing["Month"].isin(fy27_months)].copy()

# Preserve any existing actuals — merge them into the new plan rows
if not already_loaded.empty:
    print(f"Found {len(already_loaded)} existing FY2027 rows — preserving actuals.")
    actual_lookup = already_loaded.set_index(["Category", "Month"])["Actual"].to_dict()
    plan["Actual"] = plan.apply(
        lambda r: actual_lookup.get((r["Category"], r["Month"]), 0.0), axis=1
    )
    existing = existing[~existing["Month"].isin(fy27_months)]

plan["Variance"]     = plan["Actual"] - plan["Plan"]
plan["Variance_Pct"] = ((plan["Actual"] - plan["Plan"]) / plan["Plan"].replace(0, pd.NA) * 100).round(1)

# ── Align columns ──────────────────────────────────────────────────────────
for col in existing.columns:
    if col not in plan.columns:
        plan[col] = None
plan = plan[existing.columns]  # match column order exactly

# ── Append and write back ──────────────────────────────────────────────────
combined = pd.concat([existing, plan], ignore_index=True)
print(f"Combined OPEX rows: {len(combined)}")

wb = load_workbook(BOARD_DATA)
ws = wb["OPEX"]

# Clear existing sheet content and rewrite
ws.delete_rows(1, ws.max_row)

# Write header
ws.append(combined.columns.tolist())

# Write data rows
for row in combined.itertuples(index=False):
    ws.append(list(row))

wb.save(BOARD_DATA)
print("Done — TSTT_Board_Data.xlsx updated with FY2027 OPEX plan.")
print("\nAnnual plan by category (TT$M):")
for cat, total in plan.groupby("Category")["Plan"].sum().sort_values(ascending=False).items():
    print(f"  {cat:<30} TT${total/1e6:>8.2f}M")
print(f"\n  {'TOTAL':<30} TT${plan['Plan'].sum()/1e6:>8.2f}M")
