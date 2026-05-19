"""
Creates TSTT_Master_Data_Template.xlsx — a single spreadsheet containing
every data input needed by the board report dashboard, with zero duplication.

Design principles:
  - One row per natural grain (Month+Segment, Month+Product, etc.)
  - Raw values only — totals and derived metrics are computed in the dashboard
  - Actuals and AOP share the same sheet structure (distinguished by column pairs)
  - All monetary values entered as RAW TTD (e.g. 1,500,000 not 1.5 or 1,500)
  - The dashboard handles all conversion to millions/thousands for display
  - All percentages as real percentages (e.g. 2.5, not 0.025)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
CLR = {
    "sheet_header": "0D1117",   # near-black
    "tab_green":    "196127",
    "tab_blue":     "1F3864",
    "tab_purple":   "3B1F64",
    "tab_teal":     "1F5464",
    "tab_orange":   "7F3B00",
    "tab_red":      "641F1F",
    "tab_grey":     "3A3A3A",
    "hdr_green":    "196127",
    "hdr_blue":     "1F3864",
    "hdr_purple":   "3B1F64",
    "hdr_teal":     "1F5464",
    "hdr_orange":   "7F3B00",
    "hdr_red":      "641F1F",
    "hdr_grey":     "444444",
    "desc_bg":      "F2F2F2",
    "key_bg":       "E8F0FE",
    "white":        "FFFFFF",
    "light_border": "CCCCCC",
}

FONT_HDR  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_DESC = Font(name="Calibri", italic=True, color="555555", size=9)
FONT_KEY  = Font(name="Calibri", bold=True, color="1F3864", size=10)
FONT_BODY = Font(name="Calibri", size=10)
FONT_TITLE= Font(name="Calibri", bold=True, size=14, color="FFFFFF")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color=CLR["light_border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def write_sheet(ws, hdr_color, title, subtitle, columns, tab_color=None):
    """
    columns: list of (col_name, description, width, number_format)
    Returns the row number after headers (where data should start).
    """
    if tab_color:
        ws.sheet_properties.tabColor = tab_color

    # Row 1 — sheet title bar
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(columns))
    c = ws.cell(1, 1, title)
    c.font = FONT_TITLE
    c.fill = fill(hdr_color)
    c.alignment = ALIGN_L

    # Row 2 — subtitle / grain description
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=len(columns))
    c = ws.cell(2, 1, subtitle)
    c.font = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    c.fill = fill("1A1A2E")
    c.alignment = ALIGN_L

    # Row 3 — column headers
    ws.row_dimensions[3].height = 36
    for ci, (col_name, _, width, _) in enumerate(columns, 1):
        c = ws.cell(3, ci, col_name)
        c.font = FONT_HDR
        c.fill = fill(hdr_color)
        c.alignment = ALIGN_C
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Row 4 — column descriptions
    ws.row_dimensions[4].height = 32
    for ci, (_, desc, _, _) in enumerate(columns, 1):
        c = ws.cell(4, ci, desc)
        c.font = FONT_DESC
        c.fill = fill(CLR["desc_bg"])
        c.alignment = ALIGN_C
        c.border = BORDER

    # Freeze panes below description row
    ws.freeze_panes = "A5"
    return 5  # first data row


# ─────────────────────────────────────────────────────────────────────────────
# README
# ─────────────────────────────────────────────────────────────────────────────
ws_readme = wb.active
ws_readme.title = "README"
ws_readme.sheet_properties.tabColor = "000000"
ws_readme.column_dimensions["A"].width = 28
ws_readme.column_dimensions["B"].width = 90

readme_rows = [
    ("TSTT BOARD REPORT — MASTER DATA FILE", None, True, "0D1117", "FFFFFF", 20),
    ("Single source of truth for all dashboard inputs. Do not duplicate values across sheets.", None, False, "0D1117", "AAAAAA", 10),
    (None, None, False, None, None, None),
    ("GENERAL RULES", None, True, "196127", "FFFFFF", 12),
    ("Units",         "Enter monetary values exactly as they appear in the source — raw TTD (e.g. 1500000, not 1500 or 1.5). The dashboard converts to millions/thousands for display.", False, None, None, 10),
    ("Percentages",   "Enter as real % — e.g. enter 2.5 for 2.5%, not 0.025.", False, None, None, 10),
    ("Dates",         'Month column format: Mon-YY — e.g. "Apr-26", "Jan-27". Jan–Mar use the second year of the FY.', False, None, None, 10),
    ("Blanks",        "Leave cells blank (not zero) when data is not yet available. Zero means zero.", False, None, None, 10),
    ("Actuals vs AOP","Each sheet holds both Actual and AOP in paired columns (Revenue / Revenue_AOP). Enter both in the same row.", False, None, None, 10),
    ("Derived values","Do NOT enter: totals, GP (= Rev − COS), ARPU (= Rev / Subs), EBITDA Margin, variances. These are computed.", False, None, None, 10),
    (None, None, False, None, None, None),
    ("SHEET GUIDE", None, True, "1F3864", "FFFFFF", 12),
    ("P&L_Segments",      "Core P&L by business segment and month. ONE row per (Month, Segment). This drives Financial_Monthly totals.", False, None, None, 10),
    ("OPEX_Detail",       "Detailed OPEX by category and month. ONE row per (Month, Category). Must sum to Total_OPEX in P&L_Segments.", False, None, None, 10),
    ("Consumer_Products", "Consumer sub-segment detail — Prepaid, Postpaid, WTTx, TV, Other. ONE row per (Month, Product).", False, None, None, 10),
    ("Business_Products", "Business Sales product breakdown with direct costs and MRR. ONE row per (Month, Product).", False, None, None, 10),
    ("DPDI_Products",     "DPDI product revenue. COS comes from P&L_Segments.DPDI and is distributed proportionally. ONE row per (Month, Product).", False, None, None, 10),
    ("Amplia_Commercial", "Amplia commercial KPIs — ARPU, gross adds, churn. ONE row per Month.", False, None, None, 10),
    ("Cash_CAPEX",        "Cash position, net debt, free cash flow, and capital expenditure. ONE row per Month.", False, None, None, 10),
    ("Collections_Billing","Monthly billings and collections by customer group. ONE row per (Month, Group).", False, None, None, 10),
    ("AR_Aging",          "Accounts receivable aging by customer group and aging bucket. ONE row per (Month, Group, Bucket).", False, None, None, 10),
    ("Pipeline",          "Business Sales pipeline by stage. ONE row per (Month, Stage).", False, None, None, 10),
    ("Renewals",          "At-risk contract renewals. ONE row per contract renewal record.", False, None, None, 10),
    ("Prepaid_ARPU_Buckets","Prepaid subscriber distribution by ARPU bucket. ONE row per (Month, Bucket).", False, None, None, 10),
    ("Prepaid_Data_Usage","Monthly prepaid data consumption metrics. ONE row per Month.", False, None, None, 10),
    (None, None, False, None, None, None),
    ("SEGMENT NAMES (use exactly)", None, True, "3B1F64", "FFFFFF", 12),
    ("P&L Segments",   "Consumer Sales | Business Sales | DPDI | Amplia | Other", False, None, None, 10),
    ("Consumer Products", "Prepaid | Postpaid | WTTx | TV | Residential Security | Other", False, None, None, 10),
    ("Business Products", "Circuit & Data | Cloud | Carrier | Enterprise Fixed | CPE | WTTx | Security | Other", False, None, None, 10),
    ("DPDI Products",  "E-Tender | E-Health | E-Kiosk | PayPr | E-Cashbook | E-Pay | Other", False, None, None, 10),
    ("OPEX Categories","Personnel | Maintenance | Bad Debt | Advertising & PR | Consultancy & Professional Fees | Other Operating Expenses", False, None, None, 10),
    ("AR / Collections Groups", "Government | Non-Government", False, None, None, 10),
    ("AR Aging Buckets","0-30d | 31-60d | 61-90d | 90-360d | 360+d", False, None, None, 10),
    ("Pipeline Stages","Prospect | Qualified | Proposal | Negotiation | Closed Won | Closed Lost", False, None, None, 10),
    ("Risk Levels",    "Low | Medium | High | Critical", False, None, None, 10),
]

ws_readme.row_dimensions[1].height = 36
for i, row in enumerate(readme_rows, 1):
    if row[0] is None:
        continue
    label, val, bold, bg, fg, sz = row
    ws_readme.row_dimensions[i].height = 20 if not bold else 26
    ca = ws_readme.cell(i, 1, label)
    cb = ws_readme.cell(i, 2, val or "")
    if bold and bg:
        ws_readme.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ca.fill = fill(bg)
        ca.font = Font(name="Calibri", bold=True, color=fg, size=sz)
        ca.alignment = ALIGN_L
    else:
        ca.font = Font(name="Calibri", bold=True, color="333333", size=sz or 10)
        cb.font = Font(name="Calibri", color="444444", size=sz or 10)
        ca.alignment = ALIGN_L
        cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# P&L_Segments
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("P&L_Segments")
cols = [
    ("Month",         "Mon-YY format e.g. Apr-26",                              12, "@"),
    ("Segment",       "Consumer Sales | Business Sales | DPDI | Amplia | Other", 20, "@"),
    ("Revenue",       "Actual revenue (raw TTD)",                               16, "#,##0"),
    ("Revenue_AOP",   "AOP / budget revenue (raw TTD)",                         16, "#,##0"),
    ("Revenue_PY",    "Same month prior year revenue (raw TTD)",                16, "#,##0"),
    ("COS",           "Cost of sales — actual (raw TTD)",                       16, "#,##0"),
    ("COS_AOP",       "Cost of sales — AOP (raw TTD)",                          16, "#,##0"),
    ("Gross_Profit",  "DERIVED: Revenue − COS. Do not enter.",                  16, "#,##0"),
    ("GP_AOP",        "DERIVED: Revenue_AOP − COS_AOP. Do not enter.",          16, "#,##0"),
    ("OPEX",          "Total operating expenses — actual (raw TTD)",            16, "#,##0"),
    ("OPEX_AOP",      "Total operating expenses — AOP (raw TTD)",               16, "#,##0"),
    ("EBITDA",        "DERIVED: Gross_Profit − OPEX. Do not enter.",            16, "#,##0"),
    ("EBITDA_AOP",    "DERIVED: GP_AOP − OPEX_AOP. Do not enter.",              16, "#,##0"),
    ("PAT",           "Profit after tax — actual (raw TTD). Group level only.", 16, "#,##0"),
    ("PAT_AOP",       "Profit after tax — AOP (raw TTD).",                      16, "#,##0"),
    ("Bad_Debt",      "Bad debt charge included in OPEX (raw TTD)",             14, "#,##0"),
]
write_sheet(ws, CLR["hdr_green"], "P&L BY SEGMENT",
            "Grain: one row per (Month × Segment) | 5 segments × 26 months = 130 rows | Monetary: raw TTD",
            cols, tab_color="196127")

# ─────────────────────────────────────────────────────────────────────────────
# OPEX_Detail
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("OPEX_Detail")
cols = [
    ("Month",       "Mon-YY format e.g. Apr-26",                                              12, "@"),
    ("Segment",     "Consumer Sales | Business Sales | DPDI | Amplia | Other | Group Total",  20, "@"),
    ("Category",    "Personnel | Maintenance | Bad Debt | Advertising & PR | Consultancy & Professional Fees | Other Operating Expenses", 34, "@"),
    ("Actual",      "Actual OPEX spend this month (raw TTD)",      14, "#,##0"),
    ("AOP",         "Budgeted OPEX this month (raw TTD)",           14, "#,##0"),
    ("PY",          "Prior year same month OPEX (raw TTD)",         14, "#,##0"),
]
write_sheet(ws, CLR["hdr_green"], "OPEX DETAIL",
            "Grain: one row per (Month × Segment × Category) | Must sum to OPEX column in P&L_Segments | Monetary: raw TTD",
            cols, tab_color="196127")

# ─────────────────────────────────────────────────────────────────────────────
# Consumer_Products
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Consumer_Products")
cols = [
    ("Month",           "Mon-YY format e.g. Apr-26",                                      12, "@"),
    ("Product",         "Prepaid | Postpaid | WTTx | TV | Residential Security | Other",  22, "@"),
    ("Revenue",         "Actual revenue (raw TTD)",                                        14, "#,##0"),
    ("Revenue_AOP",     "AOP / budget revenue (raw TTD)",                                  14, "#,##0"),
    ("Revenue_PY",      "Same month prior year revenue (raw TTD)",                         14, "#,##0"),
    ("Subscribers",     "Closing subscriber count (units)",                                14, "#,##0"),
    ("Subscribers_AOP", "AOP subscriber target",                                           14, "#,##0"),
    ("Gross_Adds",      "New activations this month (units)",                              14, "#,##0"),
    ("Gross_Adds_AOP",  "AOP gross adds target",                                           14, "#,##0"),
    ("Churn_Pct",       "Monthly churn % of opening base (e.g. 2.5 = 2.5%)",              14, "0.00%"),
    ("ARPU",            "DERIVED: Revenue / Subscribers. Do not enter.",                   14, "#,##0.00"),
]
write_sheet(ws, CLR["hdr_blue"], "CONSUMER PRODUCTS",
            "Grain: one row per (Month × Product) | Revenue must sum to Consumer Sales in P&L_Segments | Monetary: raw TTD",
            cols, tab_color="1F3864")

# ─────────────────────────────────────────────────────────────────────────────
# Business_Products
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Business_Products")
cols = [
    ("Month",              "Mon-YY format e.g. Apr-26",                                            12, "@"),
    ("Product",            "Circuit & Data | Cloud | Carrier | Enterprise Fixed | CPE | WTTx | Security | Other", 26, "@"),
    ("Revenue",            "Actual revenue (raw TTD)",                                             14, "#,##0"),
    ("Revenue_AOP",        "AOP / budget revenue (raw TTD)",                                       14, "#,##0"),
    ("Revenue_PY",         "Same month prior year revenue (raw TTD)",                              14, "#,##0"),
    ("Direct_Costs",       "Direct cost of sales — actual (raw TTD)",                             14, "#,##0"),
    ("Direct_Costs_AOP",   "Direct cost of sales — AOP (raw TTD)",                                14, "#,##0"),
    ("Gross_Profit",       "DERIVED: Revenue − Direct_Costs. Do not enter.",                      14, "#,##0"),
    ("GP_Margin_Pct",      "DERIVED: GP / Revenue × 100. Do not enter.",                          14, "0.0%"),
    ("MRR",                "Monthly recurring revenue — actual (raw TTD)",                        14, "#,##0"),
    ("MRR_AOP",            "Monthly recurring revenue — AOP (raw TTD)",                           14, "#,##0"),
    ("Contribution",       "Contribution margin after direct costs (raw TTD)",                    14, "#,##0"),
]
write_sheet(ws, CLR["hdr_blue"], "BUSINESS SALES PRODUCTS",
            "Grain: one row per (Month × Product) | Revenue must sum to Business Sales in P&L_Segments | Monetary: raw TTD",
            cols, tab_color="1F3864")

# ─────────────────────────────────────────────────────────────────────────────
# DPDI_Products
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("DPDI_Products")
cols = [
    ("Month",          "Mon-YY format e.g. Apr-26",                              12, "@"),
    ("Product",        "E-Tender | E-Health | E-Kiosk | PayPr | E-Cashbook | E-Pay | Other", 22, "@"),
    ("Revenue",        "Actual revenue (raw TTD)",                               14, "#,##0"),
    ("Revenue_AOP",    "AOP / budget revenue (raw TTD)",                         14, "#,##0"),
    ("Revenue_PY",     "Same month prior year revenue (raw TTD)",                14, "#,##0"),
    ("Direct_Costs",   "DERIVED from P&L_Segments DPDI COS, distributed by revenue proportion. Do not enter.", 20, "#,##0"),
    ("Gross_Profit",   "DERIVED: Revenue − Direct_Costs. Do not enter.",         14, "#,##0"),
]
write_sheet(ws, CLR["hdr_purple"], "DPDI PRODUCTS",
            "Grain: one row per (Month × Product) | Revenue must sum to DPDI in P&L_Segments | COS is proportionally allocated from P&L_Segments | Monetary: raw TTD",
            cols, tab_color="3B1F64")

# ─────────────────────────────────────────────────────────────────────────────
# Amplia_Commercial
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Amplia_Commercial")
cols = [
    ("Month",                "Mon-YY format e.g. Apr-26",                   12, "@"),
    ("ARPU",                 "Average revenue per user — actual (TTD)",     14, "#,##0.00"),
    ("ARPU_AOP",             "ARPU — AOP / budget (TTD)",                   14, "#,##0.00"),
    ("Gross_Adds",           "New activations — actual (units)",            14, "#,##0"),
    ("Gross_Adds_AOP",       "New activations — AOP target (units)",        14, "#,##0"),
    ("Monthly_Churn",        "Churned subscribers this month — actual",     14, "#,##0"),
    ("Monthly_Churn_AOP",    "Churned subscribers — AOP target",            14, "#,##0"),
    ("Closing_Subscribers",  "Total active subscribers at month end",       16, "#,##0"),
    ("Net_Port",             "Net porting: port-ins minus port-outs",       14, "#,##0"),
]
write_sheet(ws, CLR["hdr_teal"], "AMPLIA COMMERCIAL KPIs",
            "Grain: one row per Month | Financial data is in P&L_Segments (Amplia row) | This sheet is commercial/operational KPIs only",
            cols, tab_color="1F5464")

# ─────────────────────────────────────────────────────────────────────────────
# Cash_CAPEX
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Cash_CAPEX")
cols = [
    ("Month",           "Mon-YY format e.g. Apr-26",                       12, "@"),
    ("Cash_Balance",    "Cash and cash equivalents at month end (raw TTD)", 16, "#,##0"),
    ("Net_Debt",        "Total debt minus cash (raw TTD). Positive = net debt.", 16, "#,##0"),
    ("FCF",             "Free cash flow for the month (raw TTD)",           14, "#,##0"),
    ("CAPEX_Actual",    "Capital expenditure spent YTD (raw TTD)",          16, "#,##0"),
    ("CAPEX_Plan",      "Annual capital expenditure budget (raw TTD)",      16, "#,##0"),
    ("CAPEX_Monthly_Actual", "CAPEX spent this month only (raw TTD)",       18, "#,##0"),
    ("CAPEX_Monthly_Plan",   "CAPEX budget for this month (raw TTD)",       16, "#,##0"),
]
write_sheet(ws, CLR["hdr_teal"], "CASH · NET DEBT · CAPEX",
            "Grain: one row per Month | Monetary: raw TTD",
            cols, tab_color="1F5464")

# ─────────────────────────────────────────────────────────────────────────────
# Collections_Billing
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Collections_Billing")
cols = [
    ("Month",            "Mon-YY format e.g. Apr-26",                  12, "@"),
    ("Group",            "Government | Non-Government | Consumer | Amplia", 22, "@"),
    ("Billings",         "Total billings issued this month (raw TTD)",  16, "#,##0"),
    ("Collections",      "Cash collected this month (raw TTD)",         16, "#,##0"),
    ("Collections_Pct",  "DERIVED: Collections / Billings × 100. Do not enter.", 16, "0.0%"),
    ("Prior_Month_Billings", "Billings from the prior month (raw TTD) — used for prior-month collection rate", 20, "#,##0"),
    ("Prior_Month_Collections_Pct", "Collections this month as % of prior month billings", 24, "0.0%"),
]
write_sheet(ws, CLR["hdr_teal"], "COLLECTIONS & BILLING",
            "Grain: one row per (Month × Group) | Monetary: raw TTD",
            cols, tab_color="1F5464")

# ─────────────────────────────────────────────────────────────────────────────
# AR_Aging
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("AR_Aging")
cols = [
    ("Month",   "Mon-YY format e.g. Apr-26. Use the reporting month.", 12, "@"),
    ("Group",   "Government | Non-Government",                          18, "@"),
    ("Bucket",  "0-30d | 31-60d | 61-90d | 90-360d | 360+d",          14, "@"),
    ("Amount",  "Outstanding receivable in this bucket (raw TTD)",      16, "#,##0"),
]
write_sheet(ws, CLR["hdr_orange"], "ACCOUNTS RECEIVABLE AGING",
            "Grain: one row per (Month × Group × Bucket) | 2 groups × 5 buckets = 10 rows per month | Monetary: raw TTD",
            cols, tab_color="7F3B00")

# ─────────────────────────────────────────────────────────────────────────────
# Pipeline
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Pipeline")
cols = [
    ("Month",         "Mon-YY format e.g. Apr-26",                                       12, "@"),
    ("Stage",         "Prospect | Qualified | Proposal | Negotiation | Closed Won | Closed Lost", 20, "@"),
    ("Segment",       "Business Sales segment this pipeline belongs to",                  18, "@"),
    ("Deal_Count",    "Number of deals in this stage",                                    12, "#,##0"),
    ("Value_TTD",     "Total value of deals in this stage (raw TTD)",                    16, "#,##0"),
    ("Avg_Deal_Size", "DERIVED: Value_TTD / Deal_Count. Do not enter.",                  16, "#,##0"),
    ("Win_Rate_Pct",  "Historical win rate for this stage (e.g. 35 = 35%)",              14, "0.0%"),
    ("Weighted_Value","DERIVED: Value_TTD × Win_Rate_Pct / 100. Do not enter.",          16, "#,##0"),
]
write_sheet(ws, CLR["hdr_orange"], "BUSINESS SALES PIPELINE",
            "Grain: one row per (Month × Stage × Segment) | Monetary: raw TTD",
            cols, tab_color="7F3B00")

# ─────────────────────────────────────────────────────────────────────────────
# Renewals
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Renewals")
cols = [
    ("Customer",         "Customer / account name",                                   24, "@"),
    ("Product_Service",  "Product or service being renewed",                           22, "@"),
    ("Segment",          "Business Sales segment",                                     18, "@"),
    ("ACV_TTD",          "Annual contract value (raw TTD)",                            16, "#,##0"),
    ("Expiry_Date",      "Contract expiry date (Mon-YY format)",                       14, "@"),
    ("Risk_Level",       "Low | Medium | High | Critical",                             12, "@"),
    ("Status",           "Active | In Negotiation | At Risk | Won | Lost",             18, "@"),
    ("Action_Plan",      "Brief description of retention action",                       30, "@"),
    ("Owner",            "Account manager or team responsible",                         18, "@"),
    ("Last_Updated",     "Date this record was last updated (Mon-YY)",                 14, "@"),
]
write_sheet(ws, CLR["hdr_red"], "AT-RISK RENEWALS",
            "Grain: one row per contract renewal record | ACV in TTD'000",
            cols, tab_color="641F1F")

# ─────────────────────────────────────────────────────────────────────────────
# Prepaid_ARPU_Buckets
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Prepaid_ARPU_Buckets")
cols = [
    ("Month",          "Mon-YY format e.g. Apr-26",                   12, "@"),
    ("ARPU_Bucket",    "ARPU range label e.g. $0-$50 | $51-$100 | $101-$200 | $201+", 20, "@"),
    ("Subscribers",    "Subscriber count in this bucket",              14, "#,##0"),
    ("Revenue",        "Revenue generated by this bucket (raw TTD)",   14, "#,##0"),
    ("Pct_of_Total",   "DERIVED: Subscribers / Total Subs × 100. Do not enter.", 14, "0.0%"),
]
write_sheet(ws, CLR["hdr_grey"], "PREPAID ARPU DISTRIBUTION",
            "Grain: one row per (Month × ARPU_Bucket) | Revenue in TTD'000",
            cols, tab_color="3A3A3A")

# ─────────────────────────────────────────────────────────────────────────────
# Prepaid_Data_Usage
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Prepaid_Data_Usage")
cols = [
    ("Month",               "Mon-YY format e.g. Apr-26",                      12, "@"),
    ("Total_GB",            "Total data consumed by prepaid base (GB)",        14, "#,##0"),
    ("Data_Users",          "Number of prepaid subs who used data this month", 14, "#,##0"),
    ("MoU_Per_User",        "Average minutes of use per prepaid subscriber",   14, "#,##0.0"),
    ("Bundle_Subscribers",  "Prepaid subs on a data bundle",                   16, "#,##0"),
    ("Bundle_Revenue",      "Revenue from data bundles (raw TTD)",             16, "#,##0"),
    ("GB_Per_User",         "DERIVED: Total_GB / Data_Users. Do not enter.",   14, "#,##0.00"),
    ("Bundle_Pct",          "DERIVED: Bundle_Subs / Total Prepaid Subs × 100. Do not enter.", 14, "0.0%"),
]
write_sheet(ws, CLR["hdr_grey"], "PREPAID DATA USAGE",
            "Grain: one row per Month | GB figures in gigabytes | Revenue in TTD'000",
            cols, tab_color="3A3A3A")


# ─────────────────────────────────────────────────────────────────────────────
# Final save
# ─────────────────────────────────────────────────────────────────────────────
out_path = "TSTT_Master_Data_Template.xlsx"
wb.save(out_path)
print(f"Created: {out_path}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
